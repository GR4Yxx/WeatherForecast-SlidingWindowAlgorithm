#This file is resposible for getting data from firebase and storing it into an excel sheet for the algorithm

import pyrebase
import time
from openpyxl import Workbook
from openpyxl import load_workbook

config = {
  "apiKey": "AIzaSyAGhNysOvktIq6-snzeRV8J6pV9YAPTN6M",
  "authDomain": "nodeweatherforecast.firebaseapp.com",
  "databaseURL": "https://nodeweatherforecast-default-rtdb.asia-southeast1.firebasedatabase.app",
  "projectId": "nodeweatherforecast",
  "storageBucket": "nodeweatherforecast.appspot.com",
  "messagingSenderId": "374473788440",
  " ": "1:374473788440:web:40c2e7b213ed9d46d5c76d"
}
firebase = pyrebase.initialize_app(config)
db=firebase.database()

wb=load_workbook('data.xlsx')
ws=wb.active

humidity=[0.0,100.0,0.0] #current,min,max
temperature=[0.0,100.0,0.0] #current,min,max

records=0

def getRows():
    global records    
    records=0
    for row in ws.iter_rows():
        records+=1
    
    print('no of records is :',records)

def getData():
    global humidity,temperature
    humidity[0]=float(db.child("humidity").get().val())
    temperature[0]=float(db.child("temperature").get().val())

    if humidity[0]<humidity[1]:
        humidity[1]=humidity[0]
    if humidity[0]>humidity[2]:
        humidity[2]=humidity[0]

    if temperature[0]<temperature[1]:
        temperature[1]=temperature[0]
    if temperature[0]>temperature[2]:
        temperature[2]=temperature[0]
    
def logData():
    getRows()
    #sr no
    ws.cell(row=records+1, column=1).value = records

    #humidity
    ws.cell(row=records+1, column=2).value = temperature[1]

    #min temp
    ws.cell(row=records+1, column=3).value = temperature[2]

    #max temp
    ws.cell(row=records+1, column=4).value = humidity[0]

    #rainfall
    ws.cell(row=records+1, column=5).value = 0



getRows()
while True:
    
    getData()
    print('-----------------------------------------------------------------------------------------------')
    print(f' (Humidity)    current:{humidity[0]}  minimum:{humidity[1]} maximum:{humidity[2]}')
    print(f' (Temperature) current:{temperature[0]}  minimum:{temperature[1]} maximum:{temperature[2]}')
    logData()
    wb.save('data.xlsx')
    time.sleep(5)


