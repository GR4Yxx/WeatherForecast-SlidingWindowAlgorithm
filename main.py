#This file handles the main algorithm that reads records and processes its

from array import *
from openpyxl import load_workbook
from scipy import spatial
from numpy import minimum,array_equal
CD=[]
PD=[]
tempWindow=[]
slidingW=[]
euclideanD=[]

minWindow=0
k=0
n=8



window_size=7
past_window_size=14
rowCounter=2
maxRowCounter=0
readings=8
wb=load_workbook('data.xlsx')
ws=wb.active

def calcED():
    global euclideanD,minWindow

    for x in range(8):
        Y =spatial.distance.cdist(CD, slidingW[x], 'euclidean')
        euclideanD.append(Y)
    for i in range(8):
        euclideanD[i]=euclideanD[i].tolist()

    minx=minimum(euclideanD[0],euclideanD[1])
    for y in range(2,8):
        minx=minimum(minx,euclideanD[y])
    for z in range(8):
        if array_equal(euclideanD[z],minx):
            minWindow=z    

def printTable(x,S):
    print(S)
    for a in x:
        print(a)

def printWindows(windows):
    print(f'Windows :{len(windows)}')
    for x in range(readings):
        printTable(windows[x],f'Window {x+1}')

def init():
    global PD,CD,rowCounter,tempWindow,slidingW
    for row in ws.iter_rows(min_row=2,min_col=2,max_row=past_window_size+1, max_col=5,  values_only=True):
        PD.append(row)
    for row in ws.iter_rows(min_row=2,min_col=8,max_row=window_size+1, max_col=11,  values_only=True):
        CD.append(row)

    while rowCounter<readings+2:
        for row in ws.iter_rows(min_row=rowCounter,min_col=2,max_row=window_size-1+rowCounter, max_col=5,  values_only=True):
            tempWindow.append(row)
        slidingW.append(tempWindow)
        rowCounter+=1
        tempWindow=[]
    
    



#Main Function
init()
#Step 1
printTable(CD,'Current Data')

#Step 2
printTable(PD,'Past Data')

#Step 3
printWindows(slidingW)

#Step 4
calcED()
minWindow=1
print(f'Least window is {minWindow}')

#Step 5




#Calculating euclidean distance
# calcED(1)